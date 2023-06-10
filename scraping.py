import datetime
import email
import time
from imaplib import IMAP4_SSL
from base64 import b64decode
from quopri import decodestring

from openpyxl.reader.excel import load_workbook
from email.header import decode_header
from openpyxl.workbook import Workbook

from config import hidden_vars
from support_func import date_out, y, check_data_in_db, _profit, write_into_db
from apscheduler.schedulers.background import BlockingScheduler

from test import f_re


def mail_connect() -> list:
    global con
    con = IMAP4_SSL("imap.mail.ru")
    con.login(hidden_vars.mailbox, hidden_vars.mail_pass)
    con.select(hidden_vars.mail_path)
    result, data_connect = con.search(None, "UNSEEN")
    msg_list = list(data_connect[0].decode('UTF-8').split())
    if msg_list:
        return msg_list


def mail_processing(msg_list: list) -> list:
    print('have unread emails')
    to_be_write_into_db = list()
    for i in msg_list:
        result, data = con.fetch(i, "(RFC822)")
        msg = email.message_from_bytes(data[0][1])
        date_time_letter = date_out(email.utils.parsedate_to_datetime(msg["Date"]))
        subject = decode_header(msg["Subject"])[0][0].decode()
        if hidden_vars.subject_keywords_xls in subject:
            to_be_write_into_db.append(date_time_letter)
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_maintype() == 'multipart' or part.get('Content-Disposition') is None:
                        continue
                    filename = part.get_filename()
                    transfer_encoding = part.get_all('Content-Transfer-Encoding')
                    if transfer_encoding and transfer_encoding[0] == 'base64':
                        filename_parts = filename.split('?')
                        filename = b64decode(filename_parts[3]).decode(filename_parts[1])
                        if '.xlsx' or '.xls' in filename:
                            with open(
                                    f'{hidden_vars.mail_path}/{filename}', 'wb'
                            ) as new_file:
                                new_file.write(part.get_payload(decode=True))
                            y.upload(f'{hidden_vars.mail_path}/{filename}',
                                     f'/shippers/Mobex/{filename}', overwrite=True)
                            to_be_write_into_db.append(filename)
                            print('Saved attachments XLSX or XLS')
        if hidden_vars.subject_keywords_apple in subject:
            to_be_write_into_db.append(date_time_letter)
            for part in msg.walk():
                if part.get_content_maintype() == 'text' and part.get_content_subtype() == 'plain':
                    transfer_encoding = part.get_all('Content-Transfer-Encoding')
                    if transfer_encoding and transfer_encoding[0] == 'base64':
                        text = b64decode(part.get_payload()).decode()
                        to_be_write_into_db = add_excel_data(text, date_time_letter)
                    if transfer_encoding and transfer_encoding[0] == 'quoted-printable':
                        text = decodestring(part.get_payload()).decode('utf-8')
                        to_be_write_into_db = add_excel_data(text, date_time_letter)
                    print('received a letter without an attachment but in a text that falls under the conditions')
        else:
            print('letters do not meet the conditions')
        return to_be_write_into_db


def add_excel_data(text_data: str, date: list) -> list:
    to_be_write_into_db = [date]
    wrong_words = ['Dyson', 'Galaxy']
    temp_ = list()
    in_list = list()
    for line in text_data.replace(' 2sim', '').split('\n'):
        result = any(item in line for item in wrong_words)
        print(result)
            #and len(line) > 5 and '-' in line:
    #             in_list.append(line.partition('-'))
    # #
    # for i in in_list:
    #     print(f"{i[0].strip()} - {f_re(i[2])}")

    # wb = Workbook()
    # ws = wb.active
    # ws.title = "Лист1"
    # ws.append(['Наименование', 'Цена', 'Заказ'])
    # for row in in_list:
    #     ws.append([row[0].strip(), int(row[2].replace(' ', ''))])
    # filename = hidden_vars.subject_keywords_apple.replace(' ', '_') + '_' + date[:10] + '.xlsx'
    # wb.save(f'{hidden_vars.mail_path}/{filename}')
    # y.upload(f'{hidden_vars.mail_path}/{filename}', f'/shippers/Mobex/{filename}', overwrite=True)
    # to_be_write_into_db.append(filename)
    # return to_be_write_into_db


def from_xls_into_db(data_list: list) -> None:
    condition = {'к.xlsx': 'optmobex_xiaomi',
                 'sams.xlsx': 'optmobex_samsung',
                 'Apple': 'optmobex_apple'
                 }
    price_list_name = str()
    for key in condition.keys():
        if key in data_list[1]:
            price_list_name = condition.get(key)
    result_checking = check_data_in_db(data_list[0], price_list_name)
    if result_checking:
        print(f'В БД {price_list_name} уже есть такой прайс:', data_list[1])
        pass
    else:
        print(f'Заношу в БД {price_list_name} такой прайс:', data_list[1])
        price_list_temp_ = list()
        wb = load_workbook(hidden_vars.mail_path + '/' + data_list[1])
        ws = wb["Лист1"]
        rows = ws.max_row
        cols = ws.max_column - 1
        for i in range(2, rows):
            string = str()
            for j in range(1, cols + 1):
                cell = ws.cell(row=i, column=j)
                string = string + str(cell.value) + ' '
            price_list_temp_.append([data_list[0], string.strip(' ').split(' ')])
        price_list = list()
        for i in price_list_temp_:
            price_list.append({'date': i[0],
                               'product': ' '.join(i[1][:-1]),
                               'enter_cost': int(i[1][-1]),
                               'out_cost': _profit(int(i[1][-1]))})
        price_list_temp_.clear()
        write_into_db(distributor_price_list=price_list_name,
                      price_list=price_list)
        print('Запись:', data_list[1], 'завершена')


def mail_parsing():
    while True:
        response = mail_connect()
        if response:
            prepare_letters = mail_processing(response)
            # if prepare_letters:
            #     from_xls_into_db(prepare_letters)
        time.sleep(120)


scheduler = BlockingScheduler()
scheduler.add_job(mail_parsing, 'cron', day_of_week='0-5', hour='11-14', minute='*')

if __name__ == '__main__':
    mail_parsing()

